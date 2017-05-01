import os
import shutil
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


def format_excel(excel_path):

    wb = load_workbook(filename=excel_path)
    
    # grab all sheets except the first readme one
    sheet_list = wb.get_sheet_names()[1:]
    
    yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    for sheet in sheet_list:
        ws = wb.get_sheet_by_name(sheet)
        
        # unbold left-most column after first two rows
        for cell in ws['A'][2:]:
            cell.font = cell.font.copy(bold=False)
            
        # format numbers to include commas and hide decimals
        for row in ws.iter_rows(row_offset=2):
            for cell in row:
                cell.number_format = '#,##0'
                
        # highlight top left cell
        ws['A1'].fill = yellowFill
        
        # adjust cell width to match column data
        adjust_cell_width(ws)
        
    wb.save(excel_path)
 
 
def cell_length(cell):
    # http://stackoverflow.com/a/40935194/4355916
    return len(str(cell.value))
    
 
def adjust_cell_width(ws):

    col_count = 0
    
    for column_cells in ws.columns:
        cell_length_list = []
        
        # if it's the first column, include the first cell when we resize column
        if col_count == 0:
            valid_cells = [c for c in column_cells if c.value]
            
        # start with the second row-- merged cells screw up cell length
        # remove cells with value of None                
        else:
            valid_cells = [c for c in column_cells[1:] if c.value]
        
        for cell in valid_cells:
            cell_length_list.append(cell_length(cell))                    
    
        # if any cell has a value
        if cell_length_list:
            length = max(cell_length_list)
            ws.column_dimensions[column_cells[0].column].width = length
            
        col_count += 1

if __name__ == '__main__':
    root_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    output_dir = os.path.join(root_dir, 'output')

    #src_file = os.path.join(output_dir, 'tree_cover_stats_2015.xlsx')
    src_file = os.path.join(output_dir, 'tree_cover_stats_2015_IRL.xlsx')
    path_to_excel = os.path.join(output_dir, 'dummy.xlsx')

    if os.path.exists(path_to_excel):
        os.remove(path_to_excel)

    shutil.copy2(src_file, path_to_excel)

    format_excel(path_to_excel)