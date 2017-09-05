import os
import shutil
import argparse
from openpyxl.styles import PatternFill, Alignment
from openpyxl import load_workbook


def main():

    parser = argparse.ArgumentParser(description='Postprocess excel output to standard format')
    parser.add_argument('--input', '-i', help='Input excel sheet to process', required=True)
    parser.add_argument('--output', '-o', help='Max admin level to summarize')
    args = parser.parse_args()
    
    format_excel(args.input, args.output)


def format_excel(excel_path, output_path):

    wb = load_workbook(filename=excel_path)
    
    # grab all sheets except the first readme one
    sheet_list = wb.get_sheet_names()[1:]
    
    yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    aligned_left = Alignment(horizontal='left')
    
    for sheet in sheet_list:
        ws = wb.get_sheet_by_name(sheet)
        
        # unbold left-most column after first two rows
        for cell in ws['A'][2:]:
            cell.font = cell.font.copy(bold=False)
            
        # left justify left-most column cells
        for cell in ws['A']:
            cell.alignment = aligned_left
            
        # left justify first two rows 
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.alignment = aligned_left
            
        # format numbers to include commas and hide decimals
        for row in ws.iter_rows(row_offset=2):
            for cell in row:
                cell.number_format = '#,##0'
                
        # highlight top left cell
        ws['A1'].fill = yellowFill
        
        # adjust cell width to match column data
        adjust_cell_width(ws)
        
    if output_path:
        wb.save(output_path)
    else:
        wb.save(excel_path)
 
 
def cell_length(cell):
    # http://stackoverflow.com/a/40935194/4355916
    try:
        length = len(str(cell.value))
    except UnicodeEncodeError:
        length = len(cell.value.encode('utf-8'))

    return length
    
 
def adjust_cell_width(ws):

    col_count = 0
    
    for column_cells in ws.columns:
        cell_length_list = []
        
        # if it's the first column, include the first cell when we resize column
        if col_count == 0:
            valid_cells = [c for c in column_cells if c.value]
            
        # start with the second row-- merged cells screw up cell length
        # remove cells with value of None                
        else:
            valid_cells = [c for c in column_cells[1:] if c.value]
        
        for cell in valid_cells:
            cell_length_list.append(cell_length(cell))                    
    
        # if any cell has a value
        if cell_length_list:
            length = max(cell_length_list)

            # add to the width of the gain column
            if 'gain' in ws.title.lower() and col_count == 1:
                length += 4
                
            # catch instances where no loss, just the year
            # keep spacing relatively even for all columns
            if length <= 14:
                length = 15
                
            ws.column_dimensions[column_cells[0].column].width = length
            
        col_count += 1

if __name__ == '__main__':
    main()
